import os
import shutil
import openpyxl

# オリジナルのファイルパスを指定
path = r'C:\Users\N030\Desktop\faceRecognition-master\faceRecognition\excel_copy'
filename = '勤怠管理.xlsx'

# 対象年
year = 2022

# 事業所
office = 'SOONESS'

# 作成する社員 noは社員番号、nameは社員名、Paid有給、Joined入社日　Section 課名 url前年度勤怠
users = [
    {'no' : 1, 'name': '山田 太郎', 'paid' : 0,  'section' : 'メンバー',   'hiredate':'2012/04/05', 'url': 'http://a' },
    {'no' : 2, 'name': '竹下 次郎', 'paid' : 7,  'section' : '職業指導員', 'hiredate':'2012/05/05', 'url': 'http://b' },
    {'no' : 3, 'name': '松下 三郎', 'paid' : 10, 'section' : 'ヘルパー' ,  'hiredate':'2012/05/10', 'url': 'http://c' },
    {'no' : 4, 'name': '本田 史郎', 'paid' : 14, 'section' : '課長' ,      'hiredate':'2012/06/07', 'url': 'http://d' },
]

for user in users :
    newfilename = "%s.勤怠管理(%s)   .xlsx" % (str(user['no']), user['name'])

    # ファイルの作成
    shutil.copyfile(path + '\\' + filename, path + '\\' + newfilename)

    # Bookの読み込み
    wb = openpyxl.load_workbook(path + '\\' + newfilename)

    # シートの読み込み
    sheet = wb['名前']

    # 社員番号
    sheet.cell(column = 2,row = 2).value = user['no']

    # 氏名
    sheet.cell(column = 2,row = 3).value = user['name']

    # 課名
    sheet.cell(column = 2,row = 4).value = user['section']

    # 事業所
    sheet.cell(column = 2,row = 5).value = office

    # 年
    sheet.cell(column = 2,row = 9).value = year

    # 年２
    sheet.cell(column = 2,row = 10).value = year + 1

    # 有給
    sheet.cell(column = 2,row = 11).value = user['paid']

    # 前年度勤怠
    sheet.cell(column = 2,row = 12).value = user['url']

    # 入社日
    sheet.cell(column = 2,row = 13).value = user['hiredate']

    # 保存先とファイル名の設定
    wb.save(path + '\\' + newfilename)
    