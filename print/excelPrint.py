# -*- coding: utf-8 -*-

import PySimpleGUI as sg
import datetime
from dateutil.relativedelta import relativedelta
from bs4 import BeautifulSoup
import requests
import json
import openpyxl
import shutil

def serverLogin(session):

    url = 'http://localhost/kintai/syains/login'

    # サーバーログイン
    with session.get(url) as response:

        # クッキー取得
        co = response.cookies

        # csrfトークンを取得
        bs = BeautifulSoup(response.text, "html.parser")
        token = bs.find('input', attrs={ 'name' : '_csrfToken' }).get('value')
        payload = {'login_id': '123', 'password' : 123}

    return session.post(url, data = payload, headers = {'X-CSRF-Token': token}, cookies = co)

def getSyains() :
    # セッション作成
    with requests.Session() as session :
		# サーバーにログイン
        with serverLogin(session) as response:
            bs = BeautifulSoup(response.text, "html.parser")
            token = bs.find('input', attrs={ 'name' : '_csrfToken' }).get('value')

            # ログイン失敗時
            if response.status_code != 200:
                sg.popup_error('認証失敗')
                return None

			# クッキー取得
            co = response.cookies

    # # 送信情報
    # payload = {'user_id': id, 'facepath': path}

    # url = 'http://localhost/kintai/syains/getSyains'
    # with session.post(url, data = payload, headers = {'X-CSRF-Token': token}, cookies = co) as response:
    #     status_code = response.status_code
    #     rst = response.text


    url = 'http://localhost/kintai/syains/getSyains'
    with session.get(url, headers = {'X-CSRF-Token': token}, cookies = co) as response:

        if response.status_code != 200:
            sg.popup_error('サーバーエラー')
            return None

        # csrfトークンを取得
        token = bs.find('input', attrs={ 'name' : '_csrfToken' }).get('value')
        # he = {'X-CSRF-Token': token}

        # クッキー取得
        # co = response.cookies

        syains = json.loads(response.text)

        list = {}

        for syain in syains:
            # print(key, ":", json_dicti[key])
            list['id:' + syain['id']] = syain['syain_name']

        return list

def createDtList() :
    # dtList = {}
    dtList = []
    dt = datetime.datetime.now()
    dt.strftime('%Y年%m月')

    flg = True
    i = 0
    dt = dt + relativedelta(months = 1)
    while flg:
        if i <=  -36: break
        dt = dt + relativedelta(months = -1)
        
        # dtList[dt.strftime('%Y%m')] =  dt.strftime('%Y年%m月')
        dtList.append(str(dt.strftime('%Y年%m月')))
        i -= 1

    return dtList

def getPrintData(values) :

    cookies = None
    with requests.Session() as session :
		# サーバーにログイン
        with serverLogin(session) as response:
            bs = BeautifulSoup(response.text, "html.parser")
            token = bs.find('input', attrs={ 'name' : '_csrfToken' }).get('value')

            # ログイン失敗時
            if response.status_code != 200:
                sg.popup_error('認証失敗')
                return None

			# クッキー取得
            cookies = response.cookies


        # 送信情報
        payload = {'date': values['date']}

        for key in values.keys():
            if 'id:' in key :
                if values[key] == True:
                    payload[key] = key[3:len(key)]

        url = 'http://localhost/kintai/facesattendances/get_kintai_data'
        with session.post(url, data = payload, headers = {'X-CSRF-Token': token}, cookies = cookies) as response:
            # status_code = response.status_code

            if response.status_code != 200:
                sg.popup_error('サーバーエラー')
                return None

            rst = response.text

    return json.loads(rst)

def editPrint(datas):
    len(datas)

    shutil.copy('./kintaikanrihyou.xlsx', './edit.xlsx')
    wb = openpyxl.load_workbook('edit.xlsx')

    maeMon = ''
    maeSyainNo = ''
    sheet = None
    dayAds = 7

    startformat = None

    for data in datas:
        if maeMon != data['kin_date'][0:-3] or maeSyainNo != data['syain_no']: 
            # 改ページを行う（ワークシートをコピー）
            sheet = wb.copy_worksheet(wb['moto'])
            sheet.title = data['kin_date'][0:-3] + data['syain_name']

            f = sheet.cell(column = 1, row = 1).number_format
            sheet.cell(column = 1, row = 1).value = data['kin_date'][0:4]
            sheet.cell(column = 1, row = 1).number_format = f

            f = sheet.cell(column = 7, row = 1).number_format
            sheet.cell(column = 7, row = 1).value = data['kin_date'][5:7]
            sheet.cell(column = 7, row = 1).number_format = f

            f = sheet.cell(column = 2, row = 4).number_format
            sheet.cell(column = 2, row = 4).value = data['syain_no']
            sheet.cell(column = 2, row = 4).number_format = f

            f = sheet.cell(column = 18, row = 4).number_format
            sheet.cell(column = 18, row = 4).value = data['syain_name']
            sheet.cell(column = 18, row = 4).number_format = f

            f = sheet.cell(column = 34, row = 4).number_format
            sheet.cell(column = 34, row = 4).value = data['section_name']
            sheet.cell(column = 34, row = 4).number_format = f

            startformat = sheet.cell(column = 4, row = dayAds + int(data['kin_date'][-2:])).number_format    # 書式が消えるので退避

        maeMon = data['kin_date'][0:-3]
        maeSyainNo = data['syain_no']

        # 始業時刻
        
        sheet.cell(column = 5, row = dayAds + int(data['kin_date'][-2:])).value = data['start_time']
        sheet.cell(column = 5, row = dayAds + int(data['kin_date'][-2:])).number_format = startformat    # 消えた書式を戻す
        # 終業時刻
        sheet.cell(column = 10, row = dayAds + int(data['kin_date'][-2:])).value = data['end_time']
        sheet.cell(column = 10, row = dayAds + int(data['kin_date'][-2:])).number_format = startformat


    wb.save('edit.xlsx')

def display():

    sg.theme('Dark Blue 3')

    name_list = list(range(50))

    syains = getSyains()

    # スクロール表示させたい画面構成を指定
    cols = [
        [sg.Text('社員名', size=(5, 1))]] + \
        [[sg.Checkbox(syain[1], key=syain[0], default=True)] for syain in syains.items()]

        # [[sg.Text(str(rrt), size=(5, 1)), sg.Input(size=(15, 1))] for rrt in name_list]

    dtList = createDtList()

    # 上記で指定した内容をsg.Columnの引数に入れる。 scrollableをTrueにする
    layout = [
        [sg.Text('対象月', size=(5, 1))],
        [[sg.Combo(dtList, default_value=dtList[0],key = 'date', size=(20,1)) ]],
        [sg.Checkbox("全チェック", default=True, key = 'zencheck', enable_events = True)],
        [sg.Column(cols, scrollable=True , vertical_scroll_only=True, size=(200, 400))],
        [sg.Button('印刷', key = 'print'), sg.Cancel(key = 'cancel')]
    ]

    # window = sg.Window(title='Name Peaks', layout = layout)
    # rrt_list = name_list
    # rrt_to_name = {} # dict

    # ウィンドウ生成
    window1 = sg.Window('印刷指示', layout)
    index = 0

    while True:
        event, values = window1.read()

        if event is None:
            break

        if event == 'zencheck':

            check = values['zencheck']

            for syain in syains:
                window1[syain].Update(value = check)
        if event == 'print':
            data = getPrintData(values)

            if data == None:
                continue

            editPrint(data)

        if event == 'cancel':
            break

    # ウィンドウ破棄
    window1.close()

# git 確認