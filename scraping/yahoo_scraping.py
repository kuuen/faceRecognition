import time
from selenium import webdriver
import os
import shutil
import openpyxl
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException

# ★導入方法★
# seleniumをインストール
#   pip install selenium
#
# https://sites.google.com/chromium.org/driver/downloads?authuser=0 にて使用しているchromeのバージョンのwebdriverをダウンロードする
# ダウンロードしたexeファイルをpython.exeと同じディレクトリに保存する（場所は環境変数で確認できる）
# 店舗情報.xlsxを実行ファイルの同ディレクトリに置く
# main() の引数に商品名を指定する

# WebDriverのインスタンスを作成
driver = webdriver.Chrome() 

# 対象ページを開く
def get(url) :
  driver.get(url)
  time.sleep(2) # 2秒待機

# Yahooショッピングを開いて商品を検索する
def main(s):
  ''' s:検索キーワード
  '''

  # 開く
  get('https://shopping.yahoo.co.jp/')

  # 検索欄にキーワードを入力
  driver.find_element_by_id("ss_yschsp").send_keys(s)

  time.sleep(2) # 2秒待機
  # 検索ボタンクリック
  driver.find_element_by_id("ss_srch_btn").click()

  linkstrs = driver.find_elements_by_class_name('_2EW-04-9Eayr')
  page = 0
  index = 0
  # 商品リンクの文字列がまったく同じものがある
  linkRireki = {}

  while True :
    page += 1
    # メインループ
    index += listLoop(page, linkstrs, linkRireki, index)

    # ページは非同期描画で書き込まれる。リンクを再読み込み
    linkstrs = driver.find_elements_by_class_name('_2EW-04-9Eayr')

    # 再読み込みしても商品数が増えていなかったら最終ページと判断できる
    if index == len(linkstrs) :
      break

  driver.quit() # ブラウザを閉じる

def listLoop(page, linkstrs, linkRireki, yomikomizumiIndex):
  """
  page      :読み込む対象のページ
  linkstrs  :対象のリンク
  linkRireki:同じ名前のリンクリスト、値は読み込み済みのindex
  yomikomizumiIndex:読込済みリンクのindex
  """

  # ページindex
  index = 0

  # 改ページ時のスキップ数
  skipcount = 0

  # i = 1
  # リンク一覧を参照
  for str in linkstrs :
    # ページのスキップ skipcount-1はスキップ完了状態
    if page > 1 and skipcount != -1:
      if (yomikomizumiIndex) > skipcount:
        skipcount += 1
        continue
      else:
        skipcount = -1

    # リンクの文字列でリンクを特定
    links = driver.find_elements_by_partial_link_text(str.text)

    # リンクが複数ある場合は工夫が必要
    if len(links) == 1 :
      link = links[0]
      linkRireki[str.text] = 0
    else:
      # 特定のリンクを指定する
      if  linkRireki.get(str.text) == None:
        linkRireki[str.text] = 0
      else:
        linkRireki[str.text] += 1

      link = links[linkRireki[str.text]]
      
    # 1行舐めたらスクロールする
    # if i % 5 == 0:
      # 下スクロールする
      # driver.find_element_by_tag_name('body').click()
      # driver.find_element_by_tag_name('body').send_keys(Keys.PAGE_DOWN)
      # scrollByElemAndOffset(link, -10)

    # i += 1

    try:
      # リンクを開く
      link.click()
    except ElementClickInterceptedException :
      # 同一商品が複数ある場合にexceptionを吐く場合がある
      print('同一商品でのエラー')
      index += 1
      continue

    time.sleep(2) # 2秒待機

    # 新しいタブに切り替える
    driver.switch_to.window(driver.window_handles[1])

    # 業者ページを参照
    referGyousya()

    # 新しいタブを閉じる
    driver.close()

    # 閉じたタブから前のタブを切り替える
    driver.switch_to.window(driver.window_handles[0])
    index += 1

  return index

# 業者ページ処理
def referGyousya() :

  link = None
  list = None
  try :
    # 通常店舗とPayPayモール店があるから処理を振り分ける
    link = driver.find_element_by_partial_link_text('会社概要')
    list = referGyousyaYahoo()
  except NoSuchElementException :
    list = referGyousyaPayPay()

  # 既に書き込みが完了した企業は対象外
  if list != None:
    # エクセル書き込み
    witeExcel(list)

# 業者ページを開くPayPayモールの場合
def referGyousyaPayPay() :

  try :
    linkstr = driver.find_element_by_class_name('ItemSeller_name')
  except NoSuchElementException:
    # ページが見つからない場合があった
    return 

  link = driver.find_element_by_partial_link_text(linkstr.text)
  # 販売元をクリック
  link.click()
  time.sleep(2) # 2秒待機

  try :
    # 会社概要・お買い物ガイドクリック
    linkstr = driver.find_element_by_class_name('StoreService_item')
  except NoSuchElementException :
    # ページが見つからない場合があった
    return

  link = driver.find_element_by_partial_link_text(linkstr.text)
  # 会社概要・お買い物ガイド
  link.click()
  time.sleep(2) # 2秒待機

  zyouhous = driver.find_elements_by_class_name('StoreInfo_item')

  list = {}
  taisyou = True
  # 情報領域全体を参照する
  for zyouhou in zyouhous:
    # 列名、値を分割する
    ss = zyouhou.text.split('\n')

    # 住所列の場合、
    if ss[0].find('会社名（商号）') > -1:
      list['companyName'] = ss[1]
    elif ss[0].find('郵便番号') > -1:
      list['postCode'] = ss[1]
    elif ss[0].find('郵便番号') > -1:
      list['postCode'] = ss[1]
    elif ss[0].find('住所') > -1:
      if 'adress1' not in list:
        list['adress1'] = ss[1]
    elif ss[0].find('所在地') > -1:
      list['adress2'] = ss[1]
    elif ss[0].find('代表者') > -1:
      list['representative'] = ss[1]
    elif ss[0].find('ストア名') > -1:
      list['shopNameKana'] = ss[1]
      list['shopName'] = ss[2]
    elif ss[0].find('ストア名') > -1:
      list['shopName'] = ss[1]
    elif ss[0].find('ストア紹介') > -1:
      list['setumei'] = ss[1]
    elif ss[0].find('関連ストア') > -1:
      list['relatedStore'] = ss[1]
    elif ss[0].find('運営責任者') > -1:
      list['operationManager'] = ss[1]
    elif ss[0].find('お問い合わせ') > -1:
      list['tel'] = ss[1]
      list['mail'] = ss[2]
    elif ss[0].find('ストア営業日') > -1:
      list['Time'] = ss[1]

  # 沖縄以外の場合は対象外
  if list['adress1'].find('沖縄') == -1 and list['adress2'].find('沖縄') == -1:
    taisyou = False
  else :
    taisyou = True

  # 対象外の場合はここで終了
  if taisyou == False:
    return None
  else :
    return list

# 業者ページを開く
def referGyousyaYahoo() :
  # 会社概要を開く
  link = driver.find_element_by_partial_link_text('会社概要')

  # 販売元をクリック
  link.click()
  time.sleep(2) # 2秒待機
  
  # 情報領域の内容を取得
  zyouhous = driver.find_elements_by_class_name('elRow')

  # ページを正しく取得できなかった場合
  if len(zyouhous) == 0:
    # 該当企業は飛ばす
    return None

  list = {}
  taisyou = True
  # 情報領域全体を参照する
  for zyouhou in zyouhous:
    # 列名、値を分割する
    ss = zyouhou.text.split('\n')

    # 各値を整理する
    if ss[0].find('住所') > -1:

      if 'adress1' not in list:
        list['adress1'] = ss[1]
      else:
        list['adress2'] = ss[1]
        
    elif ss[0].find('会社名（商号）') > -1:
      list['companyName'] = ss[1]
    elif ss[0].find('郵便番号') > -1:
      list['postCode'] = ss[1]
    elif ss[0].find('代表者') > -1:
      list['representative'] = ss[1]
    elif ss[0].find('ストア名（フリガナ）') > -1:
      list['shopNameKana'] = ss[1]
    elif ss[0].find('ストア名') > -1:
      list['shopName'] = ss[1]
    elif ss[0].find('ストア紹介') > -1:
      list['setumei'] = ss[1]
    elif ss[0].find('運営責任者') > -1:
      list['operationManager'] = ss[1]
    elif ss[0].find('電話番号') > -1:
      list['tel'] = ss[1]
    elif ss[0].find('お問い合わせファックス番号') > -1:
      list['fax'] = ss[1]
    elif ss[0].find('お問い合わせメールアドレス') > -1:
      list['mail'] = ss[1]
    elif ss[0].find('ストア営業日/時間') > -1:
      list['Time'] = ss[1]

  # 沖縄以外の場合は対象外
  if list['adress1'].find('沖縄') == -1 and list['adress2'].find('沖縄') == -1:
    taisyou = False
  else :
    taisyou = True

  # 対象外の場合はここで終了
  if taisyou == False:
    return None
  else :
    return list

def kaisyaExist(companyName, sheet) :
  ''' 企業がエクセルに既に書き込まれているかどうかのチェック
  '''

  gyouNo = 2

  result = False
  # 書き込む行番号の決定。空白行まで移動する
  while True :
    if sheet.cell(column = 2, row = gyouNo).value == None:
      result = False
      break
    if sheet.cell(column = 2, row = gyouNo).value == companyName:
      result = True
      break

    gyouNo += 1
  return result

def witeExcel(list):
  # ファイルパスを指定
  # path = r'C:\Users\N030\Desktop\faceRecognition-master\faceRecognition\scraping'
  filename = '店舗情報.xlsx'
  
  # Bookの読み込み
  wb = openpyxl.load_workbook(filename)

  # シートの読み込み
  sheet = wb['店舗情報']
  
  # 既に会社情報が載っている場合は何もしない
  if kaisyaExist(list['companyName'], sheet) :
    return

  gyouNo = 1

  # 書き込む行番号の決定。空白行まで移動する
  while True :
    gyouNo += 1

    if sheet.cell(column = 1,row = gyouNo).value == None:
      sheet.cell(column = 1,row = gyouNo).value = gyouNo - 1
      break

  # 会社名
  sheet.cell(column = 2, row = gyouNo).value = list['companyName']

  # メールアドレス
  sheet.cell(column = 3, row = gyouNo).value = list['mail']

  # 郵便番号
  sheet.cell(column = 4, row = gyouNo).value = list['postCode']

  # 住所1
  sheet.cell(column = 5, row = gyouNo).value = list['adress1']

  # 住所2
  sheet.cell(column = 6, row = gyouNo).value = list['adress2']

  # 代表者
  sheet.cell(column = 7, row = gyouNo).value = list['representative']

  # ストア名
  sheet.cell(column = 8, row = gyouNo).value = list['shopName']

  # ストア名（フリガナ）
  sheet.cell(column = 9, row = gyouNo).value = list['shopNameKana']

  # ストア紹介
  sheet.cell(column = 10, row = gyouNo).value = list['setumei']

  # 運営責任者
  sheet.cell(column = 11, row = gyouNo).value = list['operationManager']

  # 電話番号
  sheet.cell(column = 12, row = gyouNo).value = list['tel']

  # お問い合わせファックス番号 無い店舗もある
  if 'fax' in list :
    sheet.cell(column = 13, row = gyouNo).value = list['fax']

  # ストア営業日/時間
  sheet.cell(column = 14, row = gyouNo).value = list['Time']

  # 関連ストア
  if 'relatedStore' in list:  
    sheet.cell(column = 15, row = gyouNo).value = list['relatedStore']

  # ここで保存
  wb.save(filename)

    # list[]
    # print(zyouhou.text)

  # if adress.text

# 商品名称を指定する
# findItem('pixel')
# findItem('少女週末旅行 1 BD')
main('もずく')


